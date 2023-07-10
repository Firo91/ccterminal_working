from .models import Equipment, CheckHistory, EquipmentEditHistory, CustomUser
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
import pandas as pd
from django.db.models import Q
from datetime import date, datetime
from django.http import HttpResponse
import openpyxl
from .forms import EquipmentForm, CustomUserCreationForm
from django.forms.models import model_to_dict
from django.contrib.auth import  login, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import logout
from django.contrib.auth.forms import PasswordChangeForm
from .forms import CustomPasswordResetForm
from django.contrib import messages
import random
import string
import json
import pytz
from django.http import JsonResponse
from datetime import datetime, timedelta
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import user_passes_test

def group_required(group_name):
    return user_passes_test(lambda user: Group.objects.get(name=group_name) in user.groups.all())

def home(request):
    return render(request, 'home.html')

def home_bsm(request):
    return render(request, 'home_bsm.html')

@login_required
@group_required('BSM Equipment Search Users')
def bsm_equipment_search(request):
    # Get the equipment queryset
    equipments = Equipment.objects.all()
    
    # Get the selected country from the session
    selected_country = request.session.get('selected_country')
    
    # Check if a new country has been selected
    if request.GET.get('country'):
        selected_country = request.GET.get('country')
        # Save the selected country to the session
        request.session['selected_country'] = selected_country
    
    # Filter the equipment queryset based on the selected country
    if selected_country:
        equipments = equipments.filter(country=selected_country)

    query = request.GET.get('search')
    if query:
        # Filter the equipment queryset based on the search query
        equipments = equipments.filter(Q(name__icontains=query) | 
                                       Q(bax_nr__icontains=query) | 
                                       Q(tid__icontains=query) | 
                                       Q(terminal_type__icontains=query) | 
                                       Q(city__icontains=query) |
                                       Q(version__icontains=query) ) 
                                       
    
    # Render the template with the equipment queryset and selected country
    return render(request, 'equipment/bsm_equipment_search.html', {'equipments': equipments})

@login_required
def equipment_search(request):
    # Get the equipment queryset
    equipments = Equipment.objects.all()

    # Get the selected country from the session
    selected_country = request.session.get('selected_country')

    # Check if a new country has been selected
    if request.GET.get('country'):
        selected_country = request.GET.get('country')
        # Save the selected country to the session
        request.session['selected_country'] = selected_country

    # Filter the equipment queryset based on the country of the logged-in user or username matching Equipment name
    if request.user.is_authenticated:
        user_country = request.user.country
        city = request.user.city
        equipments = Equipment.objects.filter(city__icontains=city, country=user_country)
    else:
        equipments = Equipment.objects.none() # Return an empty queryset if the user is not logged in

    query = request.GET.get('search')
    if query:
        # Filter the equipment queryset based on the search query
        equipments = equipments.filter(Q(name__icontains=query) | 
                                       Q(bax_nr__icontains=query) | 
                                       Q(tid__icontains=query) | 
                                       Q(terminal_type__icontains=query) | 
                                       Q(city__icontains=query) |
                                       Q(version__icontains=query) ) 
    
    if request.method == 'POST':
        print("Entering the for loop")
        for equipment in equipments:
            has_equipment = request.POST.get(f'has_equipment_{equipment.id}')
            print(f"has_equipment_{equipment.id} value: {has_equipment}")

            if has_equipment:
                equipment.last_checked_date = date.today()
                equipment.has_equipment = True

                # Find the CustomUser instance based on the user's name
                last_user = CustomUser.objects.get(username=request.user.username)
                equipment.last_user = last_user

                equipment.save()
                print("equipment.last_checked_date:", equipment.last_checked_date)
                print("equipment.has_equipment:", equipment.has_equipment)
                print("equipment.last_user:", equipment.last_user)

                # Check if there is an existing CheckHistory entry for the same date
                existing_check_history = CheckHistory.objects.filter(
                    checked_date=datetime.now(),
                    equipment_id=equipment,
                    checked_by=request.user.username
                ).first()

                if existing_check_history:
                    # Update the existing CheckHistory entry
                    existing_check_history.checked_by = request.user.name
                    existing_check_history.equipment_name = equipment.name
                    existing_check_history.equipment_tid = equipment.tid
                    existing_check_history.checked_by_username = request.user.username
                    existing_check_history.save()
                else:
                    # Create a new CheckHistory entry
                    check_history = CheckHistory(
                        checked_by=request.user.name,
                        checked_date=datetime.now(),
                        equipment_id=equipment,
                        equipment_name=equipment.name,
                        equipment_tid=equipment.tid,
                        checked_by_username=request.user.username
                    )
                    check_history.save()

                    # Add the check history entry to the equipment's check history
                    equipment.check_history.add(check_history)

                
            
            else:
                equipment.has_equipment = False
                equipment.last_checked_date = None
                equipment.check_history.clear()

            equipment.save()

        return redirect(reverse('equipment_search') + '?' + request.GET.urlencode())

    # Render the template with the equipment queryset and selected country
    return render(request, 'equipment/equipment_search.html', {'equipments': equipments, 'selected_country': selected_country})

def upload_file(request):
    if request.method == 'POST':
        file = request.FILES['file']
        df = pd.read_excel(file)

        for _, row in df.iterrows():
            tid = row['Tid']
            equipment, created = Equipment.objects.get_or_create(tid=tid)
            
            equipment.bax_nr = row['Bax_nr']
            equipment.name = row['Name']
            equipment.terminal_type = row['Terminal_type']
            equipment.country = row['Country']
            equipment.city = row['City']
            equipment.serial_number = row['Serialnumber_terminal']
            equipment.version = row['Sw_version']
            
            # Set the date of last checked only if the equipment is newly created
            if created:
                equipment.date_last_checked = None  # or any other value representing an empty date
            
            equipment.save()

        return render(request, 'equipment/equipment_search.html')

    return render(request, 'equipment/upload_file.html')

def download_equipment_excel(request):
    # Query the Equipment objects
    equipment = Equipment.objects.all()

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Add column headings to the worksheet
    ws.append(['Bax Nr', 'Name', 'Tid', 'Terminal Type', 'Country', 'City', 'Serial Number', 'Last Checked Date', 'Last Checked By'])

    # Loop over the equipment and add each object's data to the worksheet
    for e in equipment:
        ws.append([e.bax_nr, e.name, e.tid, e.terminal_type, e.country, e.city, e.serial_number, e.last_checked_date, e.last_user.name if e.last_user else 'N/A'])

    # Set the filename for the download
    filename = "CCTerminal List.xlsx"

    # Set the content type for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response

def download_equipment_excel_history(request):
    # Query the Equipment objects
    equipment = CheckHistory.objects.all()

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Add column headings to the worksheet
    ws.append(['Name', 'Terminal ID', 'Checked By', 'Checked By (Username)', 'Checked Date'])

    # Convert the checked_date field to naive datetimes
    naive_checked_dates = [e.checked_date.replace(tzinfo=None) for e in equipment]

    # Loop over the equipment and add each object's data to the worksheet
    for e, dt in zip(equipment, naive_checked_dates):
        ws.append([e.equipment_name, e.equipment_tid, e.checked_by, e.checked_by_username, dt])

    # Set the filename for the download
    filename = "CCTerminal List History.xlsx"

    # Set the content type for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response

def download_equipment_excel_changes(request):
    # Query the Equipment objects
    equipment = EquipmentEditHistory.objects.all()

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Add column headings to the worksheet
    ws.append(['Equipment', 'Edited By', 'Edited At', 'Old Value', 'New Value'])

    # Loop over the equipment and add each object's data to the worksheet
    for e in equipment:
        ws.append([e.equipment.name, e.edited_by.username, e.edited_at, e.old_value, e.field])

    # Set the filename for the download
    filename = "CCTerminal List Changes.xlsx"

    # Set the content type for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response

def download_equipment_excel_users(request):
    # Query the Equipment objects
    equipment = CustomUser.objects.all()

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Add column headings to the worksheet
    ws.append(['Username', 'Name', 'Location', 'Country'])

    # Loop over the equipment and add each object's data to the worksheet
    for e in equipment:
        ws.append([e.username, e.name, e.city, e.country])

    # Set the filename for the download
    filename = "CCTerminal List Users.xlsx"

    # Set the content type for the response
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    # Return the response
    return response

@login_required
def edit_equipment(request, equipment_id):
    # Retrieve the equipment object
    equipment = get_object_or_404(Equipment, id=equipment_id)

    if request.method == 'POST':
        # Extract the new data from the request
        print(request.POST)  # Debugging statement
        new_data = {
            'bax_nr': request.POST.get('bax_nr'),
            'name': request.POST.get('name'),
            'tid': request.POST.get('tid'),
            'terminal_type': request.POST.get('terminal_type'),
            'country': request.POST.get('country'),
            'city': request.POST.get('city'),
            'serial_number': request.POST.get('serial_number'),
        }

     # Create edit history entries for the changed fields
        for field, new_value in new_data.items():
            old_value = getattr(equipment, field)
            if old_value != new_value:
                EquipmentEditHistory.objects.create(
                    equipment=equipment,
                    edited_by=request.user,
                    field=new_value,
                    old_value=old_value
                )

                # Update the equipment with new data
                setattr(equipment, field, new_value)

        # Save the updated equipment
        equipment.save()

        return JsonResponse({'status': 'success'})

    elif request.method == 'GET':
        # Handle GET request here if needed
        # For example, you can render the edit form template
        return render(request, 'equipment/edit_item.html', {'equipment': equipment})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
    
def equipment_changes_view(request, equipment_id):
    # Retrieve the equipment changes related to the equipment object
    equipment = Equipment.objects.get(id=equipment_id)
    equipment_changes = EquipmentEditHistory.objects.filter(equipment_id=equipment_id)

    return render(request, 'equipment_changes.html', {'equipment_changes': equipment_changes})

def check_history(request):
    check_histories = CheckHistory.objects.all().order_by('-checked_date')[:3]
    return render(request, 'equipment/check_history.html', {'check_histories': check_histories})

def custom_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Authenticate the user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # User credentials are correct, log in the user
            login(request, user)
            return redirect('equipment/equipment_search')  # Replace 'equipment_search' with the URL name of your desired page

        else:
            error_message = 'Invalid name or password. Please try again.'

    else:
        error_message = ''

    return render(request, 'equipment/login.html', {'error_message': error_message})

def custom_login_bsm(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Authenticate the user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # User credentials are correct, log in the user
            login(request, user)
            return redirect('bsm_equipment_search')  # Replace 'equipment_search' with the URL name of your desired page

        else:
            error_message = 'Invalid name or password. Please try again.'

    else:
        error_message = ''

    return render(request, 'equipment/login_bsm.html', {'error_message': error_message})

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            name = request.POST.get('name')
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            country = request.POST.get('country')
            city = request.POST.get('city')

            if password1 == password2:
                user = form.save(commit=False)
                user.username = username
                user.name = name
                user.country = country
                user.city = city
                user.set_password(password1)
                user.save()
                return redirect('custom_login')  # Redirect to the login page after successful registration
            else:
                # Handle password mismatch error
                return HttpResponse("Passwords do not match")
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'equipment/register.html', {'form': form})

def get_locations(request):
    country = request.GET.get('country')
    locations = Equipment.objects.filter(country=country).values_list('city', flat=True).distinct()
    return JsonResponse({'city': list(locations)})

def custom_logout(request):
    logout(request)
    return redirect('home')

def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Update the session to prevent the user from being logged out
            messages.success(request, 'Your password has been successfully changed.')
            return redirect('equipment_search')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'equipment/change_password.html', {'form': form})

def reset_password(request):
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            city = form.cleaned_data['city']
            
            try:
                user = CustomUser.objects.get(city=city, username=username)
            except CustomUser.DoesNotExist:
                user = None
            
            if user:
                # Generate a temporary password
                temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                
                # Update the user's password
                user.set_password(temp_password)
                user.save()
                
                messages.success(request, f'Temporary password: {temp_password}. Please login with this password and change it immediately.')
                return render(request, 'reset_password.html', {'form': form})
            else:
                messages.error(request, 'Invalid username or name.')
    else:
        form = CustomPasswordResetForm()
    
    return render(request, 'equipment/reset_password.html', {'form': form})

